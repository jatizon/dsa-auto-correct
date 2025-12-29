import os
import subprocess
import glob
import shutil
import traceback
import json
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from abc import ABC, abstractmethod


class CorrectionFailed(Exception):
    def __init__(self, message=None, aluno_path=None):
        if message is not None:
            AbstractCorrector.print_log(message, aluno_path, tipo_correcao="auto", encoding='utf-8')
            super().__init__(message)

class AbstractCorrector(ABC):
    def __init__(self, alunos_path, testcases_path, sheet_path, numero_lab, use_ai, aluno=None):
        self.alunos_path = alunos_path
        self.create_student_folders()
        self.alunos_list = sorted(os.listdir(self.alunos_path))
        self.testcases_path = testcases_path
        self.sheet_path = sheet_path
        self.numero_lab = numero_lab
        self.use_ai = use_ai
        self.aluno = aluno

        self.wb = Workbook()
        self.ws = self.wb.active

    @staticmethod
    def print_log(message, aluno_path, tipo_correcao, encoding):
        logs_correcao_path = aluno_path + f"/logs_correcao_{tipo_correcao}.txt"
        with open(logs_correcao_path, "a", encoding=encoding) as logs:
            print(message, file=logs) 

    @staticmethod
    def clear_logs_file(aluno_path, tipo_correcao):
        logs_correcao_path = aluno_path + f"/logs_correcao_{tipo_correcao}.txt"
        open(logs_correcao_path, 'w').close()

    def create_student_folders(self):
        lowercase_files = glob.glob(os.path.join(self.alunos_path, "lab*"))
        for file_path in lowercase_files:
            filename = os.path.basename(file_path)
            new_filename = filename.replace('lab', 'Lab')
            os.rename(file_path, os.path.join(self.alunos_path, new_filename))
        CPP_files = glob.glob(os.path.join(self.alunos_path, "Lab*.CPP"))
        for file_path in CPP_files:
            filename = os.path.basename(file_path)
            new_filename = filename.replace('.CPP', '.cpp')
            os.rename(file_path, os.path.join(self.alunos_path, new_filename))
        cpp_files = glob.glob(os.path.join(self.alunos_path, "Lab*.cpp"))
        for file_path in cpp_files:
            filename = os.path.basename(file_path) 
            folder_name = filename[5:].replace('.cpp','')
            new_folder_path = os.path.join(self.alunos_path, folder_name)
            os.makedirs(new_folder_path, exist_ok=True)
            shutil.move(file_path, os.path.join(new_folder_path, filename))  

    def remove_unwanted_files(self, aluno_path):
        for f in os.listdir(aluno_path):
            full_path = os.path.join(aluno_path, f)
            if f == "a.out":
                os.remove(full_path)
            elif f.endswith(".txt") and not f.startswith("logs_correcao"):
                os.remove(full_path)

    def get_and_handle_output(self, aluno_path, testcase):
        output_path = glob.glob(f'{aluno_path}/Lab*.txt')
        if not output_path:
            raise CorrectionFailed("Nao criou o arquivo txt de saida\n", aluno_path)
        output_folder = os.path.join(aluno_path, 'outputs')
        os.makedirs(output_folder, exist_ok=True)
        final_output_path = os.path.join(output_folder, f"{testcase}.txt")
        shutil.copy(output_path[0], final_output_path)
        os.remove(output_path[0])
        with open(final_output_path, encoding='latin-1') as output_file:
            return output_file.readlines()
        
    def get_student_code(self, aluno_path):
        cpp_path = glob.glob(f'{aluno_path}/Lab*.cpp')
        if not cpp_path:
            raise CorrectionFailed("Nao enviou o arquivo .cpp\n", aluno_path)
        with open(cpp_path[0], encoding='latin-1') as cpp_file:
            return cpp_file.read()

    def compile_student_code(self, aluno_path):
        if os.path.isdir(aluno_path):
            cpp_path = os.path.join(f'{aluno_path}/Lab*.cpp')
            try:
                result = subprocess.run(
                    f'gcc {cpp_path} -o {aluno_path}/a.out',
                    shell=True, 
                    check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE,
                    timeout=5
                )
                if result.returncode == 0 and result.stderr:
                    self.print_log(f"WARNINGS NA COMPILACAO:\n{result.stderr.decode('utf-8', errors='replace')}", aluno_path, tipo_correcao="auto", encoding='utf-8')
            except subprocess.CalledProcessError as e:
                raise CorrectionFailed(f"Erro de compilacao.\nCodigo de saida: {e.returncode}\n", aluno_path)
            except subprocess.TimeoutExpired:
                raise CorrectionFailed(f"Tempo limite de compilacao excedido.", aluno_path)
            except Exception as e:
                raise CorrectionFailed(f"Ocorreu um erro inesperado na compilacao: {e}\n")

    def run_student_code(self, aluno_path, testcase):
        try:
            testcase_path = os.path.join(self.testcases_path, testcase)
            subprocess.run(
                f'cp {testcase_path}/entrada* {aluno_path}',
                shell=True, 
                check=True
            )
            subprocess.run(
                './a.out', 
                cwd=aluno_path, 
                shell=True, 
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=5
            )
        except subprocess.CalledProcessError as e:
            raise CorrectionFailed(f"Erro na execucao do caso teste {testcase}.\nCodigo de saida: {e.returncode}\n", aluno_path)
        except subprocess.TimeoutExpired:
            raise CorrectionFailed(f"Tempo limite de execucao do caso teste {testcase} excedido", aluno_path)
        except Exception as e:
            raise CorrectionFailed(f"Ocorreu um erro inesperado na execucao do caso teste {testcase}\n")

    def correct_code(self, aluno_path, aluno_row, headers):
        try:
            code = self.get_student_code(aluno_path)
            if self.use_ai:
                self.clear_logs_file(aluno_path, "ia")
                response = self.do_ai_correction(aluno_path, code)
                json_str = "\n".join(response)
                response_dict = json.loads(json_str)
                self.add_corrections_to_sheet(aluno_row, response_dict, headers)
        except CorrectionFailed:
            raise CorrectionFailed()
        
    @abstractmethod
    def correct_output(self, aluno_path, testcase, output):
        pass

    @abstractmethod
    def do_ai_correction(self, aluno_path, code):
        pass

    def create_correction_sheet(self, headers):    
        wb = self.wb
        ws = self.ws
        ws.title = f"Notas_Lab{self.numero_lab}_CES11_2025"

        center_align = Alignment(horizontal="center", vertical="center")
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = center_align

        primeira_penalidade = 3  
        ultima_penalidade = len(headers) - 1 
        start_col_letter = get_column_letter(primeira_penalidade)
        end_col_letter = get_column_letter(ultima_penalidade)

        for row, aluno in enumerate(self.alunos_list, start=2):
            nome_formatado = aluno.replace("_", " ").title()
            cell = ws.cell(row=row, column=1, value=nome_formatado)
            cell.font = Font(name="Arial", size=10)
            nota_formula = f"=MAX(100+SUM({start_col_letter}{row}:{end_col_letter}{row}),0)"
            cell = ws.cell(row=row, column=2, value=nota_formula)
            cell.font = Font(name="Arial", size=10)

        for col, header in enumerate(headers, start=1):
            max_len = len(header)
            if col == 1:
                for aluno in self.alunos_list:
                    nome_formatado = aluno.replace("_", " ").title()
                    if len(nome_formatado) > max_len:
                        max_len = len(nome_formatado)
            ws.column_dimensions[get_column_letter(col)].width = max_len + 1

        last_row = len(self.alunos_list) + 3 
        cell = ws.cell(row=last_row, column=1, value="Média")
        cell.font = Font(name="Arial", size=10)
        media_formula = f"=SUM(B2:B{len(self.alunos_list)+1})/{len(self.alunos_list)}"
        cell = ws.cell(row=last_row, column=2, value=media_formula)
        cell.font = Font(name="Arial", size=10)
        
        wb.save(self.sheet_path)

    def add_corrections_to_sheet(self, aluno_row, response_dict, headers):
        wb = self.wb
        ws = self.ws
        for col_idx, header in enumerate(headers, start=1):
            if header in response_dict and response_dict[header]:
                cell = ws.cell(row=aluno_row, column=col_idx)

                cell.font = Font(name="Arial", size=10)

                if header == "Observações (sem desconto na nota)":
                    obs_text = "\n".join(f"- {err}" for err in response_dict[header])
                    existing_value = cell.value or ""
                    cell.value = f"{existing_value}\n{obs_text}".strip()
                else:
                    comment_text = "\n".join(f"- {err}" for err in response_dict[header])

                    cell.comment = Comment(comment_text, "José Alberto Feijão Tizon")

                    total_deduction = sum(
                        int(re.search(r"-\s*(\d+)", err).group(1))
                        for err in response_dict[header]
                        if re.search(r"-\s*\d+", err)
                    )
                    existing_value = cell.value if cell.value else ""
                    cell.value = f"=MAX(-{total_deduction})"
                    cell.alignment = Alignment(horizontal="right")

        wb.save(self.sheet_path)

    def make_correction(self):
        headers = [
            "", "Nota final", "Prazo", "Arquivo", "Saída", "Identação", "Bronco",
            "Índice 0", "Str. Descrição", "Agenda", "R.A.TAD", "L/E/main",
            "Global", "Busca Binária", "Func. Públicas", "fclose/free",
            "Outros", "Observações (sem desconto na nota)"
        ]
        self.create_correction_sheet(headers)
        progress = 1
        try:
            for aluno in self.alunos_list:
                if self.aluno is not None and aluno!=self.aluno:
                    continue
                print(f"Correcting... ({progress}/{len(self.alunos_list)}). Current student: {aluno}")
                aluno_path = os.path.join(self.alunos_path, aluno)
                aluno_row = progress + 1
                self.clear_logs_file(aluno_path, "auto")
                self.correct_code(aluno_path, aluno_row, headers)
                try:
                    self.compile_student_code(aluno_path)
                except CorrectionFailed:
                    continue
                finally:
                    progress += 1
                self.print_log('CORRECAO AUTOMATICA:', aluno_path, tipo_correcao="auto", encoding='utf-8')
                for testcase in os.listdir(self.testcases_path):
                    self.correct_output(aluno_path, testcase)
                self.remove_unwanted_files(aluno_path)
            print("Correction ended successfully")
        except Exception as e:
            print(f"Correction failed due to error: {e}")
            traceback.print_exc()



